def parse_selections(selections: dict):
    """Extract selected names from selections structure."""
    def _extract_selected(data):
        return [k for k, v in data.items() if v and k != "all"]

    return {
        "locations": _extract_selected(selections.get("locations", {})),
        "departments": _extract_selected(selections.get("departments", {})),
        "designations": _extract_selected(selections.get("designations", {})),
        "cost_centers": _extract_selected(selections.get("costCenters", {})),
    }
import io
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from fastapi import HTTPException
from fastapi.responses import StreamingResponse


def create_excel_template(columns: list, filename: str):
    """Generate a blank Excel template with given columns."""
    df = pd.DataFrame({col: [] for col in columns})
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def read_excel_file(file) -> tuple[list[str], list[dict]]:
    """Reads Excel content and returns headers + list of row dicts."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")

    contents = file.file.read()
    workbook = load_workbook(io.BytesIO(contents))
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    rows = [
        dict(zip(headers, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

    return headers, rows
def format_employee_name(first_name: str, last_name: str) -> str:
    """Return clean full name."""
    return f"{first_name} {last_name or ''}".strip()

# utils/excel_utils.py
import io, datetime
from openpyxl import Workbook, load_workbook
from fastapi import HTTPException
from fastapi.responses import StreamingResponse


def clean_header(h: str) -> str:
    if not h:
        return ""
    return "".join(c for c in str(h).strip().lower() if c.isalnum() or c == "_")


def clean_cell(val):
    if val is None:
        return None
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    return val


def create_employee_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = [
        "employee_code", "first_name", "last_name", "email", "mobile", "date_of_joining",
        "location_name", "department_name", "designation_name", "cost_center_name",
        "shift_policy_name", "week_off_policy_name", "notes"
    ]
    ws.append(headers)

    ws.append([
        "E001", "John", "Doe", "john.doe@example.com", "9876543210",
        datetime.date(2023, 1, 15), "Hyderabad", "Product Development Team",
        "Associate Software Engineer", "Software BU", "General Policy",
        "Hyderabad Week Off Policy", "Sample note"
    ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"}
    )


def load_excel(file_bytes: bytes):
    stream = io.BytesIO(file_bytes)
    try:
        wb = load_workbook(stream, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")
    return wb.active

# utils/salary_utils.py
def calculate_total_deduction(gi: float, gratuity: float) -> float:
    return (gi or 0) + (gratuity or 0)

# utils/salary_utils.py
def calculate_gross_salary(basic: float, hra: float, sa: float, mda: float, ca: float, ta: float) -> float:
    return sum([basic or 0, hra or 0, sa or 0, mda or 0, ca or 0, ta or 0])


# utils/work_profile_utils.py

def format_employee_name(first_name: str, last_name: Optional[str]) -> str:
    """Combine first and last name neatly."""
    return f"{first_name} {last_name or ''}".strip()
